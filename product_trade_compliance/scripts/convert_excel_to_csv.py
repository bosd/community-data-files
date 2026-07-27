#!/usr/bin/env python3
# Copyright 2024 Odoo Community Association (OCA)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).
# pylint: disable=print-used
"""Convert EU Excel file to CSV for Odoo import."""

import csv
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Input/output
excel_file = (
    "List of dual use items in Annex I to "
    "Regulation (EU) 2021/821 (September 2024).xlsx"
)
csv_file = "../data/trade_compliance_code_eu.csv"

print(f"Converting {excel_file} to CSV...")

wb = openpyxl.load_workbook(excel_file, data_only=True)
sheet = wb.active

# Find header row
header_row = None
for idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
    if row and "CODE" in [str(cell).upper() if cell else "" for cell in row]:
        header_row = idx
        headers = [str(cell).strip() if cell else "" for cell in row]
        break

if not header_row:
    print("Error: Could not find header row")
    sys.exit(1)

# Get column indices
id_idx = headers.index("ID")
code_idx = headers.index("CODE")
label_idx = headers.index("LABEL")

# Write CSV
with open(csv_file, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    # Odoo CSV header: id, name, regime_id/id, description
    writer.writerow(["id", "name", "regime_id/id", "description"])

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[code_idx]:
            continue

        row_id = str(row[id_idx]).strip() if row[id_idx] else ""
        code = str(row[code_idx]).strip()
        label = str(row[label_idx]).strip() if row[label_idx] else ""

        if code and code != "None" and row_id and row_id != "None":
            # CSV format: xmlid, code, regime_xmlid, description
            xml_id = f"compliance_code_eu_{row_id}"
            writer.writerow(
                [xml_id, code, "product_trade_compliance.regime_eu_dual", label]
            )

wb.close()
print(f"✓ Converted to {csv_file}")
print("  Ready for Odoo import!")
